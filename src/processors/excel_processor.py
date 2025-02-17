"""Обработчик Excel файлов."""

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell import Cell

from utils import ConfigManager
from utils.constants import DeliveryType


class ExcelProcessor:
    """Обработчик Excel файлов."""

    def __init__(self, file_path: Path, config_manager: ConfigManager) -> None:
        """Инициализация обработчика Excel файлов.

        Args:
            file_path: Путь к файлу
            config_manager: Менеджер конфигурации
        """
        self.file_path = file_path
        self.config_manager = config_manager
        self.df: pd.DataFrame | None = None

    def _get_cell_value(self, cell: Cell) -> str:
        """Получение значения ячейки с учетом её типа.

        Args:
            cell: Ячейка Excel

        Returns:
            str: Значение ячейки в виде строки
        """
        value = cell.value

        if cell.data_type == "n" and value is not None:
            value = str(int(value))
        elif hasattr(cell, "number_format") and "General" not in cell.number_format:
            try:
                value = cell.internal_value
            except Exception:
                pass

            value = str(value).strip() if value is not None else ""

        return str(value) if value is not None else ""

    def _get_first_word(self, value: str) -> str:
        """Получение первого слова из строки.

        Args:
            value: Исходная строка

        Returns:
            str: Первое слово или пустая строка
        """
        if not value:
            return ""
        return value.split()[0] if value.strip() else ""

    def read_data(self) -> pd.DataFrame:
        """Чтение данных из Excel файла с учетом конфигурации колонок."""
        try:
            # Получаем настройки Excel
            excel_settings = self.config_manager.settings.get_excel_settings()
            use_columns = excel_settings.get("use_columns", True)
            start_row = excel_settings.get("start_row", 1)

            # Получаем конфигурацию колонок
            columns_config = self.config_manager.columns.get_config()

            # Читаем файл
            wb = load_workbook(
                self.file_path, data_only=True
            )  # Если файл не существует, openpyxl сам выбросит ошибку
            ws = wb.worksheets[excel_settings.get("sheet_index", 0)]

            # Читаем данные
            data = []
            for row in ws.iter_rows(min_row=start_row + 1):
                row_data = {}

                # Обрабатываем общие колонки
                common_config = columns_config.get("common_columns", {})
                if "delivery" in common_config:
                    delivery_config = common_config["delivery"]
                    if "method" in delivery_config:
                        method_config = delivery_config["method"]
                        col_letter = method_config["excel_column"]
                        col_idx = self._get_column_index(col_letter)
                        cell = row[col_idx]
                        row_data["delivery.method"] = self._get_cell_value(cell)

                # Обрабатываем postal_columns
                postal_config = columns_config.get("postal_columns", {})
                for field_name, field_config in postal_config.items():
                    col_letter = field_config["excel_column"]
                    col_idx = self._get_column_index(col_letter)
                    cell = row[col_idx]
                    row_data[f"postal.{field_name}"] = self._get_cell_value(cell)

                # Обрабатываем pickpoint_columns
                pickpoint_config = columns_config.get("pickpoint_columns", {})

                # Получатель
                if "receiver" in pickpoint_config:
                    receiver_config = pickpoint_config["receiver"]
                    if "type" in receiver_config:
                        type_config = receiver_config["type"]
                        col_letter = type_config["excel_column"]
                        col_idx = self._get_column_index(col_letter)
                        cell = row[col_idx]
                        row_data["receiver.type"] = self._get_cell_value(cell)

                # Данные получателя "Лично я"
                if "self" in pickpoint_config:
                    self_config = pickpoint_config["self"]
                    for field, field_config in self_config.items():
                        col_letter = field_config["excel_column"]
                        col_idx = self._get_column_index(col_letter)
                        cell = row[col_idx]
                        row_data[f"self.{field}"] = self._get_cell_value(cell)

                # Данные другого получателя
                if "other" in pickpoint_config:
                    other_config = pickpoint_config["other"]
                    for field, field_config in other_config.items():
                        col_letter = field_config["excel_column"]
                        col_idx = self._get_column_index(col_letter)
                        cell = row[col_idx]
                        row_data[f"other.{field}"] = self._get_cell_value(cell)

                # Метод доставки (с первым словом)
                if "delivery" in pickpoint_config:
                    delivery_config = pickpoint_config["delivery"]
                    if "method" in delivery_config:
                        method_config = delivery_config["method"]
                        col_letter = method_config["excel_column"]
                        col_idx = self._get_column_index(col_letter)
                        cell = row[col_idx]
                        value = self._get_cell_value(cell)
                        if method_config.get("first_word", False):
                            value = self._get_first_word(value)
                        row_data["pickpoint.delivery.method"] = value

                data.append(row_data)

            self.df = pd.DataFrame(data)
            return self.df

        except FileNotFoundError:
            raise FileNotFoundError(f"Файл не найден: {self.file_path}")
        except Exception as e:
            raise ValueError(f"Ошибка чтения файла: {str(e)}") from e

    def _get_column_index(self, col_letter: str) -> int:
        """Получение индекса колонки из буквенного обозначения."""
        return (
            sum(
                (ord(c) - ord("A") + 1) * (26**i)
                for i, c in enumerate(reversed(col_letter))
            )
            - 1
        )
