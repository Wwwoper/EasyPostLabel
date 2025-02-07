"""Модуль для обработки альтернативных способов доставки."""

import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Font, NamedStyle, Side
from openpyxl.worksheet.worksheet import Worksheet

from processors.delivery_strategies.base import DeliveryStrategy
from utils.config import ConfigManager


class AlternativeDeliveryStrategy(DeliveryStrategy):
    """Стратегия обработки альтернативной доставки."""

    SENDER_NAME = "Беренёвой Ольги"
    SENDER_PHONE = "+79159833971"

    def __init__(self, config_manager: ConfigManager) -> None:
        """Инициализация стратегии.

        Args:
            config_manager: Менеджер конфигурации
        """
        self.config = config_manager.config

    def _get_first_word(self, text: str) -> str:
        """Получение первого слова из строки.

        Args:
            text: Исходная строка

        Returns:
            str: Первое слово
        """
        if not isinstance(text, str) or not text.strip():
            return ""
        return text.split()[0]

    def _get_first_word_right(self, text: str) -> str:
        """Получение первого слова из строки справа.

        Args:
            text: Исходная строка

        Returns:
            str: Первое слово справа
        """
        if not isinstance(text, str) or not text.strip():
            return ""
        return text.split()[0].title()

    def _combine_name(self, row: pd.Series) -> str:
        """Объединение фамилии и имени.

        Args:
            row: Строка DataFrame

        Returns:
            str: Полное имя с заглавными буквами
        """
        try:
            # Определяем тип получателя
            receiver_type = row.get("Получатель заказа", "").strip()

            if receiver_type == "Лично я":
                surname = str(row.get("Фамилия", "")).strip()
                name = str(row.get("Имя", "")).strip()
            else:  # "Другой человек"
                surname = str(row.get("Фамилия получателя заказа", "")).strip()
                name = str(row.get("Имя получателя заказа", "")).strip()

            # Проверяем наличие данных
            if not surname or not name:
                print(
                    f"\nПропуск записи: отсутствует фамилия или имя для {receiver_type}"
                )
                return ""

            # Объединяем фамилию и имя, гарантируя Title Case для каждого слова
            full_name = f"{surname} {name}".title()
            return full_name

        except Exception as e:
            print(f"\nОшибка при формировании имени: {str(e)}")
            return ""

    def _create_styles(self, workbook: Workbook) -> tuple[NamedStyle, NamedStyle]:
        """Создание стилей для Excel.

        Args:
            workbook: Рабочая книга Excel

        Returns:
            tuple: Стили для бирок и списка
        """
        birka_style = NamedStyle(name="birka")
        birka_style.font = Font(name="Century", size=13)
        border = Side(style="thin", color="000000")
        birka_style.border = Border(
            left=border, top=border, right=border, bottom=border
        )

        list_style = NamedStyle(name="list")
        list_style.font = Font(name="Century", size=12)

        workbook.add_named_style(birka_style)
        workbook.add_named_style(list_style)

        return birka_style, list_style

    def _create_labels(self, data: pd.DataFrame, sheet: Worksheet) -> None:
        """Создание бирок на листе.

        Args:
            data: DataFrame с данными
            sheet: Лист Excel для бирок
        """
        b = 1
        for idx, row in enumerate(data.itertuples(), 1):
            column = {0: "A", 1: "C", 2: "E"}[(idx - 1) % 3]

            for i, value in enumerate(
                [row.ФИО, f"в ЦВ {row.Центр_Выдачи}", f"от {self.SENDER_NAME}"]
            ):
                cell = f"{column}{b+i}"
                sheet[cell].value = value
                sheet[cell].style = "birka"

            if idx % 3 == 0:
                b += 4

        max_row = sheet.max_row
        for row in range(1, max_row + 1):
            sheet.row_dimensions[row].height = 25 if row % 4 != 0 else 3

    def _create_delivery_list(self, data: pd.DataFrame, sheet: Worksheet) -> None:
        """Создание списка доставки на листе.

        Args:
            data: DataFrame с данными
            sheet: Лист Excel для списка доставки
        """
        sheet["B1"].value = f"Заказы от {self.SENDER_NAME} {self.SENDER_PHONE}"
        sheet["B1"].style = "list"
        sheet["B2"].value = ""

        today = datetime.date.today().strftime("%d.%m.%Y")
        sheet["B3"].value = today
        sheet["C3"].value = f"{len(data)} человек в списке"

        sorted_data = data.sort_values(["Центр_Выдачи", "ФИО"])
        border_bottom = Border(bottom=Side(style="thin"))

        current_center: str | None = None
        row_num = 4

        for _, row in sorted_data.iterrows():
            sheet[f"B{row_num}"].value = row["ФИО"]
            sheet[f"C{row_num}"].value = f"в ЦВ {row['Центр_Выдачи']}"

            sheet[f"B{row_num}"].style = "list"
            sheet[f"C{row_num}"].style = "list"

            # Добавляем разделительную линию при смене центра
            new_center = str(row["Центр_Выдачи"])

            # Проверяем смену центра
            needs_border = False
            if current_center:
                needs_border = current_center != new_center

            if needs_border:
                prev_row = row_num - 1
                sheet[f"B{prev_row}"].border = border_bottom
                sheet[f"C{prev_row}"].border = border_bottom

            current_center = new_center
            sheet.row_dimensions[row_num].height = 15
            row_num += 1

        # Добавляем линию после последней записи
        if row_num > 4:
            last_row = row_num - 1
            sheet[f"B{last_row}"].border = border_bottom
            sheet[f"C{last_row}"].border = border_bottom

    def process(self, data: pd.DataFrame) -> pd.DataFrame:
        """Обработка данных об альтернативной доставке.

        Args:
            data: DataFrame с данными клиентов

        Returns:
            DataFrame: Обработанные данные
        """
        try:
            if data.empty:
                return pd.DataFrame()

            print("\nДанные для обработки:")
            print(data.head())

            # Проверяем наличие необходимых столбцов
            required_columns = [
                "Получатель заказа",
                "Список",
                "Фамилия",
                "Имя",
                "Фамилия получателя заказа",
                "Имя получателя заказа",
            ]

            missing_columns = [
                col for col in required_columns if col not in data.columns
            ]
            if missing_columns:
                raise ValueError(f"Отсутствуют столбцы: {missing_columns}")

            processed_data = data.copy()

            # Формируем ФИО и получаем центр выдачи
            processed_data["ФИО"] = processed_data.apply(self._combine_name, axis=1)
            processed_data["Центр_Выдачи"] = processed_data["Список"].apply(
                lambda x: str(x).split()[0] if pd.notna(x) else ""
            )

            # Удаляем записи с пустыми значениями
            result = processed_data[["ФИО", "Центр_Выдачи"]].copy()
            result = result[result["ФИО"].notna() & (result["ФИО"] != "")]
            result = result.drop_duplicates()

            print("\nОбработанные данные:")
            print(result)

            return result

        except Exception as e:
            print(f"\nОшибка при обработке данных: {str(e)}")
            raise

    def save(self, data: pd.DataFrame, output_path: str) -> None:
        """Сохранение результатов в файл.

        Args:
            data: DataFrame с данными клиентов
            output_path: Путь для сохранения файла
        """
        if data.empty:
            return

        wb = Workbook()
        self._create_styles(wb)

        labels_sheet = wb.active
        labels_sheet.title = "Бирки"

        delivery_sheet = wb.create_sheet("Список доставки")

        labels_sheet.column_dimensions["A"].width = 30
        labels_sheet.column_dimensions["B"].width = 1
        labels_sheet.column_dimensions["C"].width = 30
        labels_sheet.column_dimensions["D"].width = 1
        labels_sheet.column_dimensions["E"].width = 30
        labels_sheet.column_dimensions["F"].width = 1

        delivery_sheet.column_dimensions["A"].width = 5
        delivery_sheet.column_dimensions["B"].width = 30
        delivery_sheet.column_dimensions["C"].width = 30

        self._create_labels(data, labels_sheet)
        self._create_delivery_list(data, delivery_sheet)

        wb.save(output_path)
        print(f"\nДанные сохранены в файл: {output_path}")
        print(f"Сохранено {len(data)} уникальных записей")
