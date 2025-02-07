"""Модуль для обработки альтернативных способов доставки."""

import datetime
import logging
from dataclasses import dataclass
from typing import Dict, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from processors.delivery_strategies.base import DeliveryStrategy
from utils.config import ConfigManager

logger = logging.getLogger(__name__)


@dataclass
class PickupPointClient:
    """Структура данных клиента пункта выдачи."""

    full_name: str
    pickup_point: str
    receiver_type: str

    @classmethod
    def from_row(cls, row: pd.Series, config: Dict) -> Optional["PickupPointClient"]:
        """Создание клиента из строки данных."""
        try:
            receiver_type = str(row.get("Получатель заказа", "")).strip()
            logger.debug("Обработка получателя типа: %s", receiver_type)

            # Получаем конфигурацию для типа получателя
            receiver_config = None
            for r_type in config["receiver_types"].values():
                if r_type["type"] == receiver_type:
                    receiver_config = r_type
                    break

            if not receiver_config:
                logger.warning("Неизвестный тип получателя: %s", receiver_type)
                return None

            # Получаем имя и фамилию
            surname = str(
                row.get(receiver_config["name_fields"]["surname"], "")
            ).strip()
            name = str(row.get(receiver_config["name_fields"]["name"], "")).strip()
            logger.debug("Получены данные: фамилия='%s', имя='%s'", surname, name)

            if not surname or not name:
                logger.warning(
                    "Пропуск: нет имени для %s (фамилия='%s', имя='%s')",
                    receiver_type,
                    surname,
                    name,
                )
                return None

            # Получаем код пункта выдачи
            pickup_config = config["pickup_point"]
            pickup_raw = str(row.get(pickup_config["source_column"], "")).strip()
            logger.debug("Сырые данные пункта выдачи: %s", pickup_raw)

            try:
                pickup_parts = pickup_raw.split(pickup_config["split_by"])
                pickup_point = pickup_parts[pickup_config["position"]]
                logger.debug(
                    "Разбор пункта выдачи: части=%s, выбрано='%s'",
                    pickup_parts,
                    pickup_point,
                )
            except (IndexError, AttributeError) as e:
                logger.error("Ошибка разбора пункта выдачи: %s", str(e))
                return None

            if not pickup_point:
                logger.warning("Пропуск: нет центра выдачи в строке '%s'", pickup_raw)
                return None

            client = cls(
                full_name=f"{surname} {name}".title(),
                pickup_point=pickup_point,
                receiver_type=receiver_type,
            )
            logger.debug("Создан клиент: %s", client)
            return client

        except Exception as e:
            logger.exception("Ошибка создания клиента: %s", str(e))
            return None

    def __str__(self) -> str:
        """Строковое представление клиента."""
        return f"ФИО='{self.full_name}', ПВ='{self.pickup_point}', Тип='{self.receiver_type}'"


class PickupPointStrategy(DeliveryStrategy):
    """Стратегия обработки данных для пунктов выдачи."""

    def __init__(self, config_manager: ConfigManager) -> None:
        """Инициализация стратегии."""
        self.config_manager = config_manager
        self.config = config_manager.config["delivery_methods"]["types"]["Центр_Выдачи"]
        # Получаем данные отправителя из конфига
        sender_config = self.config_manager.get_setting("excel_settings.sender", {})
        self.sender_name = sender_config.get("name", "")
        self.sender_phone = sender_config.get("phone", "")

    def _create_named_style(
        self, name: str, workbook: Workbook, style_type: str = "default"
    ) -> NamedStyle:
        """Создание именованного стиля с настройками из конфига.

        Args:
            name: Имя стиля
            workbook: Рабочая книга Excel
            style_type: Тип стиля ('label' или 'list')

        Returns:
            NamedStyle: Созданный стиль
        """
        # Получаем настройки стилей
        styles = self.config_manager.get_setting(f"excel_settings.{style_type}", {})
        defaults = self.config_manager.get_setting("excel_settings.defaults", {})

        # Фиксированные значения по умолчанию
        DEFAULT_VALUES = {
            "font_name": "Century",
            "font_size": 13,
            "font_bold": False,
        }

        # Создаем стиль
        style = NamedStyle(name=name)
        style.font = Font(
            name=styles.get(
                "font_name", defaults.get("font_name", DEFAULT_VALUES["font_name"])
            ),
            size=styles.get(
                "font_size", defaults.get("font_size", DEFAULT_VALUES["font_size"])
            ),
            bold=styles.get(
                "font_bold", defaults.get("font_bold", DEFAULT_VALUES["font_bold"])
            ),
        )

        # Добавляем границы для бирок
        if style_type == "label":
            border = Side(style="thin", color="000000")
            style.border = Border(left=border, right=border, top=border, bottom=border)

        # Настройки выравнивания
        alignment = styles.get("alignment", {})
        style.alignment = Alignment(
            horizontal=alignment.get("horizontal", "left"),
            vertical=alignment.get("vertical", "center"),
            wrap_text=alignment.get("wrap_text", True),
        )

        # Добавляем стиль в книгу, если его еще нет
        if name not in workbook.named_styles:
            workbook.add_named_style(style)

        return style

    def _create_styles(self, workbook: Workbook) -> NamedStyle:
        """Создание стилей для бирок."""
        # Создаем стиль для бирок
        style = self._create_named_style("birka", workbook, "label")
        return style

    def _create_labels(
        self, data: pd.DataFrame, sheet: Worksheet, style: NamedStyle
    ) -> None:
        """Создание бирок на листе."""
        logger.info("Начало создания бирок для %d клиентов", len(data))

        col_map = {0: "A", 1: "C", 2: "E"}

        # Используем reset_index для гарантии порядка
        data = data.reset_index(drop=True)

        # Начальная строка
        current_row = 1

        # Итерируемся по индексам для сохранения порядка
        for idx in range(len(data)):
            client = data.iloc[idx]
            position = idx % 3  # Позиция в текущей строке (0, 1, 2)
            col = col_map[position]

            logger.debug(
                "Создание бирки #%d: позиция=%s%d, клиент=%s",
                idx + 1,
                col,
                current_row,
                client["ФИО"],
            )

            # Добавляем данные и применяем стили
            cells = [
                (f"{col}{current_row}", client["ФИО"]),
                (f"{col}{current_row+1}", f"в ЦВ {client['Центр_Выдачи']}"),
                (
                    f"{col}{current_row+2}",
                    f"от {self.sender_name if self.sender_name else 'Беренёвой Ольги'}",
                ),
            ]

            for cell_ref, value in cells:
                cell = sheet[cell_ref]
                cell.value = value
                cell.style = style
                logger.debug("Ячейка %s: значение='%s'", cell_ref, value)

            # Переходим к следующей группе бирок
            if position == 2:  # После третьей бирки (индексы 0,1,2)
                current_row += 4
                logger.debug("Переход на новую строку: %d", current_row)

        logger.info("Создание бирок завершено")

    def _format_worksheet(self, sheet: Worksheet) -> None:
        """Форматирование листа с бирками."""
        logger.debug("Начало форматирования листа")

        # Получаем настройки стилей
        styles = self.config_manager.get_setting("excel_settings.styling", {})
        defaults = self.config_manager.get_setting("excel_settings.defaults", {})
        spacing = styles.get("spacing", {})

        # Фиксированные значения по умолчанию, если нет в конфиге
        DEFAULT_VALUES = {
            "column_width": 30.0,
            "spacing_width": 0.5,  # Стандартная ширина промежуточных колонок
            "row_height": 25.0,
            "spacing_height": 2.5,
            "font_name": "Century",
            "font_size": 13,
            "font_bold": False,
        }

        # Получаем размеры из конфигурации
        column_width = styles.get(
            "column_width", defaults.get("column_width", DEFAULT_VALUES["column_width"])
        )
        spacing_width = spacing.get("column", {}).get(
            "width", defaults.get("spacing_width", DEFAULT_VALUES["spacing_width"])
        )
        row_height = styles.get(
            "row_height", defaults.get("row_height", DEFAULT_VALUES["row_height"])
        )
        spacing_height = spacing.get("row", {}).get(
            "height", defaults.get("spacing_height", DEFAULT_VALUES["spacing_height"])
        )

        # Устанавливаем размеры основных колонок
        for col in ["A", "C", "E"]:
            sheet.column_dimensions[col].width = column_width
            logger.debug("Установлена ширина колонки %s: %d", col, column_width)

        # Устанавливаем минимальную ширину промежуточных колонок
        for col in ["B", "D"]:
            sheet.column_dimensions[col].width = spacing_width
            logger.debug(
                "Установлена минимальная ширина колонки %s: %f", col, spacing_width
            )

        # Устанавливаем высоту строк
        for row in range(1, sheet.max_row + 1):
            height = row_height if row % 4 != 0 else spacing_height
            sheet.row_dimensions[row].height = height
            logger.debug("Установлена высота строки %d: %f", row, height)

        logger.debug("Форматирование листа завершено")

    def _create_delivery_list(
        self,
        data: pd.DataFrame,
        sheet: Worksheet,
        style: NamedStyle,
    ) -> None:
        """Создание списка доставки на листе."""
        logger.debug("Начало создания списка доставки")

        # Получаем настройки стилей
        styles = self.config_manager.get_setting("excel_settings.list", {})
        defaults = self.config_manager.get_setting("excel_settings.defaults", {})

        # Фиксированные значения по умолчанию
        DEFAULT_VALUES = {
            "column_width": 30.0,
            "row_height": 25.0,
        }

        # Создаем стиль для списка
        list_style = self._create_named_style("delivery_list", sheet.parent, "list")

        # Заголовок
        sheet["B1"].value = f"Заказы от {self.sender_name} {self.sender_phone}"
        sheet["B1"].style = list_style
        sheet["B2"].value = ""

        # Дата и количество
        today = datetime.date.today().strftime("%d.%m.%Y")
        sheet["B3"].value = today
        sheet["C3"].value = f"{len(data)} человек в списке"
        sheet["B3"].style = list_style
        sheet["C3"].style = list_style

        # Устанавливаем размеры колонок
        sheet.column_dimensions["A"].width = 5
        column_width = styles.get(
            "column_width", defaults.get("column_width", DEFAULT_VALUES["column_width"])
        )
        sheet.column_dimensions["B"].width = column_width
        sheet.column_dimensions["C"].width = column_width

        # Сортируем данные
        sorted_data = data.sort_values(["Центр_Выдачи", "ФИО"])
        border_bottom = Border(bottom=Side(style="thin"))

        current_center = None
        row_num = 4

        # Создаем список
        for _, row in sorted_data.iterrows():
            sheet[f"B{row_num}"].value = row["ФИО"]
            sheet[f"C{row_num}"].value = f"в ЦВ {row['Центр_Выдачи']}"

            sheet[f"B{row_num}"].style = list_style
            sheet[f"C{row_num}"].style = list_style

            new_center = str(row["Центр_Выдачи"])
            needs_border = current_center and current_center != new_center

            if needs_border:
                prev_row = row_num - 1
                border = Border(bottom=Side(style="thin"))
                sheet[f"B{prev_row}"].border = border
                sheet[f"C{prev_row}"].border = border

            current_center = new_center
            row_height = styles.get(
                "row_height", defaults.get("row_height", DEFAULT_VALUES["row_height"])
            )
            sheet.row_dimensions[row_num].height = row_height
            row_num += 1

        # Добавляем границу последней строке
        if row_num > 4:
            last_row = row_num - 1
            sheet[f"B{last_row}"].border = border_bottom
            sheet[f"C{last_row}"].border = border_bottom

        logger.debug("Создание списка доставки завершено")

    def _get_sort_key(self, text: str) -> tuple:
        """Получение ключа для сортировки русских букв.

        Args:
            text: Текст для сортировки

        Returns:
            tuple: Кортеж чисел для сортировки
        """
        # Словарь для русских букв
        ru_alphabet = {
            "а": 1,
            "б": 2,
            "в": 3,
            "г": 4,
            "д": 5,
            "е": 6,
            "ё": 7,
            "ж": 8,
            "з": 9,
            "и": 10,
            "й": 11,
            "к": 12,
            "л": 13,
            "м": 14,
            "н": 15,
            "о": 16,
            "п": 17,
            "р": 18,
            "с": 19,
            "т": 20,
            "у": 21,
            "ф": 22,
            "х": 23,
            "ц": 24,
            "ч": 25,
            "ш": 26,
            "щ": 27,
            "ъ": 28,
            "ы": 29,
            "ь": 30,
            "э": 31,
            "ю": 32,
            "я": 33,
            " ": 0,  # пробел имеет наименьший приоритет
        }

        # Приводим к нижнему регистру и преобразуем каждую букву в числовой код
        # Для отсутствующих в словаре символов используем их коды + 1000
        # чтобы они шли после русских букв
        return tuple(ru_alphabet.get(c, ord(c) + 1000) for c in text.lower())

    def process(self, data: pd.DataFrame) -> pd.DataFrame:
        """Обработка данных для пунктов выдачи."""
        try:
            if data.empty:
                return pd.DataFrame()

            logger.debug("Начало обработки данных пунктов выдачи")

            # Проверяем наличие всех нужных колонок
            required_columns = self.config["columns"]["required"]
            missing_columns = [
                col for col in required_columns if col not in data.columns
            ]
            if missing_columns:
                raise ValueError(f"Отсутствуют колонки: {', '.join(missing_columns)}")

            # Создаем список клиентов
            clients = []
            for _, row in data.iterrows():
                client = PickupPointClient.from_row(row, self.config)
                if client:
                    clients.append(client)

            # Сортируем клиентов по ФИО
            clients.sort(key=lambda x: x.full_name)
            logger.debug("Клиенты отсортированы по ФИО")

            # Создаем DataFrame
            result = pd.DataFrame(
                [
                    {
                        field["name"]: getattr(client, field["source"])
                        for field in self.config["columns"]["output"]
                    }
                    for client in clients
                ]
            )

            return result.drop_duplicates()

        except Exception as e:
            logger.exception("Ошибка при обработке данных: %s", str(e))
            raise

    def save(self, data: pd.DataFrame, output_path: str) -> None:
        """Сохранение данных в Excel файл."""
        # Создаем копию данных
        data_copy = data.copy()

        # Создаем новый Excel файл
        wb = Workbook()

        # Создаем листы
        labels_sheet = wb.active
        labels_sheet.title = "Бирки"
        delivery_sheet = wb.create_sheet("Список доставки")

        # Создаем стили для бирок и списка
        style = self._create_styles(wb)

        # Создаем бирки (с сортировкой по ФИО)
        labels_data = data_copy.sort_values("ФИО").copy()
        self._create_labels(labels_data, labels_sheet, style)

        # Форматируем лист с бирками
        self._format_worksheet(labels_sheet)

        # Создаем список доставки на втором листе (с сортировкой по центрам и ФИО)
        delivery_data = data_copy.sort_values(["Центр_Выдачи", "ФИО"]).copy()
        self._create_delivery_list(delivery_data, delivery_sheet, style)

        # Сохраняем файл
        wb.save(output_path)

    def _add_label(
        self, sheet, row: int, col: int, record: pd.Series, styles: dict
    ) -> None:
        """Добавление одной бирки на лист."""
        try:
            style = sheet.parent._named_styles.get("birka")
        except (AttributeError, KeyError):
            style = None

        if not style:
            raise ValueError("Style 'birka' not found")

        # Добавляем данные и применяем стили
        cells = [
            (row, record["ФИО"]),
            (row + 1, f"в ЦВ {record['Центр_Выдачи']}"),
            (row + 2, f"от {self.sender_name}"),
        ]

        for row_idx, value in cells:
            cell = sheet.cell(row=row_idx, column=col)
            cell.value = value
            cell.style = style

        # Устанавливаем ширину только для основных колонок
        if col % 2 != 0:  # только нечетные колонки (A, C, E)
            sheet.column_dimensions[get_column_letter(col)].width = styles.get(
                "column_width", 30
            )
