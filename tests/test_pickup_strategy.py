"""Тесты для стратегии обработки пунктов выдачи."""

import os
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from processors.delivery_strategies.pickup_point import PickupPointStrategy
from utils.config import ConfigManager


@pytest.fixture(autouse=True)
def setup_test_env():
    """Фикстура для настройки тестового окружения."""
    os.environ["TESTING"] = "1"
    yield
    os.environ.pop("TESTING", None)


@pytest.fixture
def config():
    """Фикстура для получения тестовой конфигурации."""
    test_config_path = Path(__file__).parent / "data" / "test_config.yaml"
    return ConfigManager(str(test_config_path))


@pytest.fixture
def pickup_strategy(config):
    """Фикстура стратегии обработки пунктов выдачи."""
    return PickupPointStrategy(config)


@pytest.fixture
def test_data():
    """Фикстура тестовых данных."""
    return pd.DataFrame(
        {
            "ФИО": [
                "Алексеев Василий",
                "Алексеев Владимир",
                "Алексеев Павел",
                "Васильев Владимир",
                "Васильев Михаил",
                "Васильев Павел",
            ],
            "Центр_Выдачи": [
                "Солнышко",
                "Бемби",
                "Бемби",
                "Глобус",
                "Солнышко",
                "Атмосфера",
            ],
        }
    )


def test_label_positions(pickup_strategy, test_data, tmp_path):
    """Тест правильности размещения бирок на листе."""
    # Создаем тестовый файл
    output_path = tmp_path / "test_labels.xlsx"
    pickup_strategy.save(test_data, str(output_path))

    # Проверяем созданный файл
    wb = load_workbook(output_path)
    sheet = wb["Бирки"]

    # Проверяем первую строку (3 бирки)
    # Первая бирка
    assert sheet["A1"].value == "Алексеев Василий"
    assert sheet["A2"].value == "в ЦВ Солнышко"
    assert sheet["A3"].value == f"от {pickup_strategy.sender_name}"

    # Вторая бирка
    assert sheet["C1"].value == "Алексеев Владимир"
    assert sheet["C2"].value == "в ЦВ Бемби"
    assert sheet["C3"].value == f"от {pickup_strategy.sender_name}"

    # Третья бирка
    assert sheet["E1"].value == "Алексеев Павел"
    assert sheet["E2"].value == "в ЦВ Бемби"
    assert sheet["E3"].value == f"от {pickup_strategy.sender_name}"

    # Проверяем вторую строку (2 бирки)
    assert sheet["A5"].value == "Васильев Владимир"
    assert sheet["A6"].value == "в ЦВ Глобус"
    assert sheet["A7"].value == f"от {pickup_strategy.sender_name}"

    assert sheet["C5"].value == "Васильев Михаил"
    assert sheet["C6"].value == "в ЦВ Солнышко"
    assert sheet["C7"].value == f"от {pickup_strategy.sender_name}"

    # Проверяем промежуточные колонки
    assert sheet.column_dimensions["B"].width == 0.5
    assert sheet.column_dimensions["D"].width == 0.5  # Такая же ширина как у колонки B

    # Проверяем высоту строк
    assert (
        sheet.row_dimensions[4].height == 2.5
    )  # промежуток между группами бирок (из конфига)


def test_label_sorting(pickup_strategy, tmp_path):
    """Тест сортировки бирок по алфавиту."""
    # Создаем неотсортированные данные
    unsorted_data = pd.DataFrame(
        {
            "ФИО": [
                "Васильев Сергей",
                "Алексеев Иван",
                "Дмитриев Павел",
                "Борисов Петр",
                "Григорьев Андрей",
            ],
            "Центр_Выдачи": ["Глобус", "Бемби", "Солнышко", "Атмосфера", "Радуга"],
        }
    )

    # Сохраняем и проверяем
    output_path = tmp_path / "test_sorted_labels.xlsx"
    pickup_strategy.save(unsorted_data, str(output_path))

    wb = load_workbook(output_path)
    sheet = wb["Бирки"]

    # Проверяем что ФИО идут в алфавитном порядке
    expected_order = [
        "Алексеев Иван",
        "Борисов Петр",
        "Васильев Сергей",
        "Григорьев Андрей",
        "Дмитриев Павел",
    ]

    # Проверяем первую строку
    assert sheet["A1"].value == expected_order[0]
    assert sheet["C1"].value == expected_order[1]
    assert sheet["E1"].value == expected_order[2]

    # Проверяем вторую строку
    assert sheet["A5"].value == expected_order[3]
    assert sheet["C5"].value == expected_order[4]


def test_label_style(pickup_strategy, test_data, tmp_path):
    """Тест стилей бирок."""
    output_path = tmp_path / "test_style_labels.xlsx"
    pickup_strategy.save(test_data, str(output_path))

    wb = load_workbook(output_path)
    sheet = wb["Бирки"]

    # Проверяем стиль первой бирки
    cell = sheet["A1"]
    assert cell.font.name == "Century"
    assert cell.font.size == 13
    assert cell.font.bold is False
    assert cell.alignment.horizontal == "center"
    assert cell.alignment.vertical == "center"
    assert cell.alignment.wrap_text is True
    assert cell.border.left.style == "thin"
    assert cell.border.right.style == "thin"
    assert cell.border.top.style == "thin"
    assert cell.border.bottom.style == "thin"

    # Проверяем данные отправителя в бирках
    assert sheet["A3"].value == f"от {pickup_strategy.sender_name}"
    assert pickup_strategy.sender_name == "Беренёвой Ольги"
    assert pickup_strategy.sender_phone == "+79159833971"


def test_pickup_strategy(pickup_strategy):
    # Ваш тестовый код
    pass


def test_delivery_list(pickup_strategy, test_data, tmp_path):
    """Тест создания списка доставки."""
    # Создаем тестовый файл
    output_path = tmp_path / "test_delivery_list.xlsx"
    pickup_strategy.save(test_data, str(output_path))

    # Проверяем созданный файл
    wb = load_workbook(output_path)
    sheet = wb["Список доставки"]

    # Проверяем заголовок
    assert (
        sheet["B1"].value
        == f"Заказы от {pickup_strategy.sender_name} {pickup_strategy.sender_phone}"
    )

    # Проверяем информацию о количестве
    assert sheet["C3"].value == "6 человек в списке"

    # Проверяем сортировку по центрам выдачи и ФИО внутри центра
    expected_data = [
        ("Васильев Павел", "в ЦВ Атмосфера"),
        ("Алексеев Владимир", "в ЦВ Бемби"),
        ("Алексеев Павел", "в ЦВ Бемби"),
        ("Васильев Владимир", "в ЦВ Глобус"),
        ("Алексеев Василий", "в ЦВ Солнышко"),
        ("Васильев Михаил", "в ЦВ Солнышко"),
    ]

    # Проверяем данные и стили
    for idx, (name, center) in enumerate(expected_data, start=4):
        # Проверяем данные
        assert sheet[f"B{idx}"].value == name
        assert sheet[f"C{idx}"].value == center

        # Проверяем стили
        for col in ["B", "C"]:
            cell = sheet[f"{col}{idx}"]
            assert cell.font.name == "Century"
            assert cell.font.size == 13
            assert cell.font.bold is False
            assert cell.alignment.horizontal == "left"
            assert cell.alignment.vertical == "center"
            assert cell.alignment.wrap_text is True

    # Проверяем размеры колонок
    assert sheet.column_dimensions["A"].width == 5
    assert sheet.column_dimensions["B"].width == 30
    assert sheet.column_dimensions["C"].width == 30

    # Проверяем высоту строк
    assert sheet.row_dimensions[4].height == 25  # стандартная высота строки

    # Проверяем границы между разными центрами
    centers = [data[1].split()[-1] for data in expected_data]
    for idx, (current, next_center) in enumerate(zip(centers, centers[1:]), start=4):
        if current != next_center:
            for col in ["B", "C"]:
                cell = sheet[f"{col}{idx}"]
                assert (
                    cell.border and cell.border.bottom
                ), f"Нет нижней границы у ячейки {col}{idx}"
                assert (
                    cell.border.bottom.style == "thin"
                ), f"Неверный стиль границы у ячейки {col}{idx}"
