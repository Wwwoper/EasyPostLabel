"""Тесты для альтернативной стратегии доставки."""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from processors.delivery_strategies.alternative import AlternativeDeliveryStrategy
from utils.config import ConfigManager


@pytest.fixture
def strategy() -> AlternativeDeliveryStrategy:
    """Фикстура для создания стратегии."""
    config_manager = ConfigManager()
    return AlternativeDeliveryStrategy(config_manager)


@pytest.fixture
def sample_data() -> pd.DataFrame:
    """Фикстура с тестовыми данными."""
    return pd.DataFrame(
        {
            "Кто будет получать заказ": ["Лично я", "Другой человек", "Лично я"],
            "Фамилия": ["иванов-петров", "Сидоров", "КУЗНЕЦОВ"],
            "Имя": ["иван", None, "ПЕТР"],
            "Фамилия получателя заказа": [None, "Петров-Иванов", None],
            "Имя получателя заказа": [None, "Иван", None],
            "Список": ["Магнит Доставка", "Пятерочка Центр", "Магнит Север"],
        }
    )


def test_get_first_word(strategy: AlternativeDeliveryStrategy) -> None:
    """Тест получения первого слова из строки."""
    assert strategy._get_first_word("Магнит Доставка") == "Магнит"
    assert strategy._get_first_word("Пятерочка") == "Пятерочка"
    assert strategy._get_first_word("") == ""
    assert strategy._get_first_word(None) == ""


def test_get_first_word_right(strategy: AlternativeDeliveryStrategy) -> None:
    """Тест получения первого слова с правильным регистром."""
    assert strategy._get_first_word_right("иванов-петров") == "Иванов-Петров"
    assert strategy._get_first_word_right("СИДОРОВ") == "Сидоров"
    assert strategy._get_first_word_right("") == ""
    assert strategy._get_first_word_right(None) == ""


def test_combine_name(
    strategy: AlternativeDeliveryStrategy, sample_data: pd.DataFrame
) -> None:
    """Тест формирования полного имени."""
    row1 = sample_data.iloc[0]
    row2 = sample_data.iloc[1]

    assert strategy._combine_name(row1) == "Иванов-Петров Иван"
    assert strategy._combine_name(row2) == "Петров-Иванов Иван"


def test_process_data(
    strategy: AlternativeDeliveryStrategy, sample_data: pd.DataFrame
) -> None:
    """Тест обработки данных."""
    result = strategy.process(sample_data)

    assert len(result) == 3
    assert list(result.columns) == ["ФИО", "Центр_Выдачи"]

    # Проверяем правильность обработки имен и центров выдачи
    assert "Иванов-Петров Иван" in result["ФИО"].values
    assert "Магнит" in result["Центр_Выдачи"].values


def test_save_output_format(
    strategy: AlternativeDeliveryStrategy, sample_data: pd.DataFrame, tmp_path: Path
) -> None:
    """Тест формата выходного файла."""
    output_path = tmp_path / "test_output.xlsx"

    # Обрабатываем и сохраняем данные
    processed_data = strategy.process(sample_data)
    strategy.save(processed_data, str(output_path))

    # Проверяем созданный файл
    wb = load_workbook(output_path)

    # Проверяем наличие листов
    assert "Бирки" in wb.sheetnames
    assert "Список доставки" in wb.sheetnames

    # Проверяем стили
    assert "birka" in wb.named_styles
    assert "list" in wb.named_styles

    # Проверяем размеры столбцов на листе бирок
    labels_sheet = wb["Бирки"]
    assert labels_sheet.column_dimensions["A"].width == 30
    assert labels_sheet.column_dimensions["B"].width == 1

    # Проверяем форматирование списка доставки
    delivery_sheet = wb["Список доставки"]
    assert "Заказы от" in delivery_sheet["B1"].value
    assert delivery_sheet["B3"].value is not None  # Дата
